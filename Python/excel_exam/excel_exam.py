# -*- coding: utf-8 -*-
# 2020.03.04 엑셀 파일 읽고/쓰기 예제
# 참고자료: https://tkdrms568.tistory.com/163

# 패키지 import
import openpyxl

# 1. 엑셀 파일을 읽기
# 1.1 파일 열기
file_path = 'input_file.xlsx'
rb = load_owrkbook(file_path)

# 1.2 sheet 읽기
sheet_name = 'sheet1'
ws = rb[sheet_name]

# 1.3 원하는 cell의 값 출력
print(ws['C4'].value)


# 2. 엑셀 파일로 쓰기
# 2.1 파일 생성
wb = openpyxl.Workbook()

# 2.2.1 sheet 추가
sheet1 = wb.active

# 2.2.2 sheet 생성
sheet_name = 'test_sheet'
sheet2 = wb.create_sheet(sheet_name)

# 2.2.3 sheet 이름 변경
sheet2.title = 'test_sheet2'

# 2.3 sheet에 값 추가
# 2.3.1 특정 cell에 값 추가
sheet2.cell(row = 3, column = 3).value = '3.3'

# 2.3.2 셀의 첫 번째 행부터 차례대로 입력
sheet2.append([1, 2, 3, 4])

# 2.4 엑셀 파일 저장
file_path = 'output_file.xlsx'
wb.save(file_path)

